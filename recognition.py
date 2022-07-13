from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
from tkinter import messagebox as msg
from openpyxl import Workbook
import cv2
import numpy as np 
import os
import sqlite3
import datetime
from datetime import date,datetime
import time
import pandas as pd
import csv



class Recognition:
    def __init__(self,root):
        self.root = root
        self.root.geometry("1540x850")
        self.root.title("Face Recognition System")

        # title
        title_lbl = Label(self.root, text="FACE RECOGNITION",font=("times new roman",35,"bold"),bg="white",fg="blue")
        title_lbl.place(x=0,y=0,width=1530,height=55)

        # img_top = Image.open("C:/Users/adoma/Desktop/another_/opencv_face_recognition-master/Tkinter/Images/train1.jpg")
        # img_top_resize = img_top.resize((1530,130),Image.ANTIALIAS)
        # self.photoimg_top = PhotoImage(img_top_resize)

        # f_lbl = Label(self.root,image=self.photoimg_top)
        # f_lbl.place(x=5,y=0,width=720,height=130)

        # 1ST top label
        img_top = Image.open('./Images/face_recog_3.png')
        img_top = img_top.resize((700,700), Image.ANTIALIAS)
        self.photoimg_top = ImageTk.PhotoImage(img_top)

        bg_img = Label(self.root, image=self.photoimg_top)
        bg_img.place(x=0,y=55,width=700,height=700)

       
        # 2ND buttom label
        img_bottom = Image.open('./Images/main_top4.jpg')
        img_bottom = img_bottom.resize((700,700), Image.ANTIALIAS)
        self.photoimg_bottom = ImageTk.PhotoImage(img_bottom)

        bg_img = Label(self.root, image=self.photoimg_bottom)
        bg_img.place(x=700,y=55,width=700,height=700)


        # # button
        b1_1 = Button(self.root,text='Facial Recognition',cursor='hand2',command=self.face_recognition, font=("times new roman",20,"bold"),bg="red",fg="white")
        b1_1.place(x=500,y=570,width=400,height=50)


    def mark_attendance(self,i,n,d):
        with open("attendance.csv","r+") as f:
            myDataList = f.readlines()
            name_list=[]
            for line in myDataList:
                entry=line.split((","))
                name_list.append(entry[0])

            #to avoid repeat attendance
            if( (i not in name_list) and (n not in name_list) and (d not in name_list) ):
                now=datetime.now()
                d1=now.strftime("%d/%m/%y")
                dtString=now.strftime("%H:%M:%S")
                f.writelines(f"\n{i},{n},{d},{dtString},{d1},Present")


    #--------- Recognition codes-------------------
    def face_recognition(self):
        # connect to database
        conn = sqlite3.connect('Database.db')
        c = conn.cursor()

        c.execute("SELECT * FROM StudentDetails")
        print(c.fetchall())

        # checking path of traingingData.yl.
        fname = "recognizer/trainingData.yml"
        if not os.path.isfile(fname):
            print("Please train the data first")
            exit(0)

        face_cascade = cv2.CascadeClassifier('./haarcascade_frontalface_alt.xml')
        cap = cv2.VideoCapture(0)
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read(fname) 

        # workbook
        # wb = Workbook()
        # ws = wb.active
        # ws.title = "ATTENDANCE-SHEET" 
        # ws['A1'] = 'INDEX-NUMBER'
        # ws['B1'] = 'FULL-NAME'
        # ws['C1'] = 'PROGRAMME'
        # ws['D1'] = 'ATTENDANCE'
        # ws['E1'] = "DATE"
        # ws['F1'] = 'TIME'   

        

        while True:
            ret, img = cap.read()
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 3)
            for (x,y,w,h) in faces:
                cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,0),3)
                ids,conf = recognizer.predict(gray[y:y+h,x:x+w])
                c.execute("SELECT name FROM StudentDetails WHERE  index_no = (?)", (ids,))
                # result1 = c.fetchone()
                # result1 = "+".join(result1)
                n = c.fetchone()
                n = "+".join(n)

                c.execute("SELECT programme FROM StudentDetails WHERE  index_no = (?)", (ids,))
                # result2 = c.fetchone()
                # result2 = "+".join(result2)
                d = c.fetchone()
                d = "+".join(d)

                c.execute("SELECT index_no FROM StudentDetails WHERE  index_no = (?)", (ids,))
                # result3 = c.fetchone()
                i = c.fetchone()
                # result3 = "+" .join(str(result3)).replace('+',' ')
                i = "+" .join(str(i)).replace("+", " ")
                # result4 = result3.replace(',', ' ').strip()
                i = i.replace(',', ' ').strip()
                # name = result[0][0]
                # print(name,result)
                print(ids,conf)
                if conf < 50:
                    # cv2.putText(img, f"Name: {result1}", (x,y-100), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (150,255,0),2)
                    cv2.putText(img, f"Name: {n}", (x,y-100), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (150,255,0),2)
                    # cv2.putText(img, f"Programme: {result2}", (x,y-65), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (150,255,0),2)
                    cv2.putText(img, f"Programme: {d}", (x,y-65), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (150,255,0),2)
                    # cv2.putText(img, f"Index_No: {result4}", (x,y-30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (150,255,0),2)
                    cv2.putText(img, f"Index_No: {i}", (x,y-30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (150,255,0),2)
                    print(ids,conf)     
                    self.mark_attendance(i, n, d)    

                    # # creating date and time
                    # ts = time.time()
                    # date = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y')
                    # timeStamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M')
                    # data = [result4, result1,result2,'PRESENT',date,timeStamp]
                    # ws.append(data)

                    # save the worksheet
                    # wb.save("attendanceNew.xlsx")
                    # ddd
            
                    # remove drop_duplicates
                    #read the saved excelsheet
                    # data = pd.read_excel("attendanceNew.xlsx")
                    
                    # df = pd.DataFrame(data)
                    # print(df)

                    # drop-duplicates
                    # new = data.drop_duplicates(subset=["INDEX-NUMBER"],keep="first")
                    # print(new)

                    # create excel writer object
                    # writer = pd.ExcelWriter('newdata.xlsx')

                    # write dataframe to excel
                    # new.to_excel(writer)

                    
                    # save the excel
                    # writer.save()

                    print('DataFrame is written successfully to Excel File.')
                else:
                    cv2.putText(img, "No Match", (x,y), cv2.FONT_HERSHEY_SIMPLEX, 2, (0,0,255),2)
            cv2.imshow("Face Recognizer", img)

            if cv2.waitKey(10) == ord('q'):
                break
        cap.release()
        cv2.destroyAllWindows()            

if __name__ == '__main__':
    root = Tk()
    obj = Recognition(root)
    root.mainloop()